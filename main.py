import logging
import re
from datetime import datetime
import pytz
import asyncio
from telegram import Update
from telegram.ext import ApplicationBuilder, CommandHandler, MessageHandler, filters, ContextTypes, Application
import pandas as pd
from openpyxl.styles import PatternFill, Alignment
from openpyxl import Workbook
from io import BytesIO
import os

# إعداد السجل - مستوى DEBUG لتتبع المشكلات
logging.basicConfig(format='%(asctime)s - %(name)s - %(levelname)s - %(message)s', level=logging.DEBUG)
logger = logging.getLogger(__name__)

# قاموس لتخزين الملاحظات المجمعة لكل مجموعة
group_notes = {}
group_keys = {}
start_time_data = {}

# قاموس لتخزين الملاحظات المرتبة لفرزها (P و H) مع message_id
sorted_notes = {}  # {chat_id: {'P': {note: {key_id: set(message_ids)}}, 'H': {note: {key_id: set(message_ids)}}}}

# قاموس لتخزين بيانات HH و P لكل محادثة
chat_data = {}

# الكلمات التي يجب تجاهلها
ignore_words = [
    "LM", "lm", "OTDR", "otdr",
    "Site", "site", "سايت",
]

# إضافة معرف المستخدم الخاص بك
YOUR_USER_ID = "6600885643"  # استبدل هذا بمعرفك العددي

# دالة update_notes مع تقييد الحجم إلى 1000 ملاحظة
def update_notes(chat_id, caption, message_id):
    logger.debug(f"Updating notes for chat_id: {chat_id}, message_id: {message_id}, caption: {caption}")
    if chat_id not in group_notes:
        group_notes[chat_id] = {}
        group_keys[chat_id] = {"P": set(), "H": set()}

    if len(group_notes[chat_id]) > 100000000000000000000000000000000:
        oldest_key = next(iter(group_notes[chat_id]))
        del group_notes[chat_id][oldest_key]
        group_keys[chat_id][oldest_key[0]].discard(oldest_key)
        logger.debug(f"Removed oldest key {oldest_key} due to size limit")

    special_match = re.match(r'\b([Hh]\d+\s[Cc]\d+)\b', caption)
    if special_match:
        keys = special_match.group(1).upper()
        notes_content = caption

        for old_key in list(group_notes[chat_id]):
            if message_id in group_notes[chat_id][old_key]:
                del group_notes[chat_id][old_key][message_id]
                if not group_notes[chat_id][old_key]:
                    del group_notes[chat_id][old_key]
                    if old_key[0] in group_keys[chat_id]:
                        group_keys[chat_id][old_key[0]].discard(old_key)
                logger.debug(f"Removed message_id {message_id} from old_key {old_key}")

        if keys in group_notes[chat_id]:
            group_notes[chat_id][keys][message_id] = notes_content
        else:
            group_notes[chat_id][keys] = {message_id: notes_content}
        group_keys[chat_id]["H"].add(keys)
        return

    matches = re.findall(r'^(?:[^\n]*?\b([Pp]\d+|[Hh]\d+|[Cc]\d+)\b)', caption)
    keys = " ".join(matches).upper()

    notes_content = caption

    for old_key in list(group_notes[chat_id]):
        if message_id in group_notes[chat_id][old_key]:
            del group_notes[chat_id][old_key][message_id]
            if not group_notes[chat_id][old_key]:
                del group_notes[chat_id][old_key]
                if old_key[0] in group_keys[chat_id]:
                    group_keys[chat_id][old_key[0]].discard(old_key)
            logger.debug(f"Removed message_id {message_id} from old_key {old_key}")

    if keys:
        main_keys = [key for key in keys.split() if not key.startswith('C')]
        c_keys = [key for key in keys.split() if key.startswith('C')]
        c_string = " ".join(c_keys)

        for sub_key in main_keys:
            sub_key = sub_key.strip()
            if sub_key:
                full_key = f"{sub_key} {c_string}".strip()

                if full_key in group_notes[chat_id]:
                    if message_id not in group_notes[chat_id][full_key]:
                        existing_notes = "\n".join(dict.fromkeys(group_notes[chat_id][full_key].values()))
                        notes_content = f"{existing_notes}\n{notes_content}"
                    group_notes[chat_id][full_key][message_id] = notes_content
                else:
                    group_notes[chat_id][full_key] = {message_id: notes_content}

                if sub_key.startswith('H'):
                    group_keys[chat_id]["H"].add(full_key)
                elif sub_key.startswith('P'):
                    group_keys[chat_id]["P"].add(full_key)
    elif notes_content:
        if "NOTES" not in group_notes[chat_id]:
            group_notes[chat_id]["NOTES"] = {}
        group_notes[chat_id]["NOTES"][message_id] = notes_content

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    logger.debug(f"Start command received for chat_id: {update.message.chat_id}")
    await update.message.reply_text('البوت الآن يعمل! أرسل الصور مع التسمية التوضيحية وسأقوم بجمع الملاحظات.')

async def stop(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    chat_id = update.message.chat_id
    logger.debug(f"Stop command received for chat_id: {chat_id}")
    group_notes.pop(chat_id, None)
    group_keys.pop(chat_id, None)
    start_time_data.pop(chat_id, None)
    chat_data.pop(chat_id, None)
    sorted_notes.pop(chat_id, None)
    await update.message.reply_text('تم إيقاف البوت لهذه المجموعة وإعادة تعيين الملاحظات.')

async def reset(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    chat_id = update.message.chat_id
    logger.debug(f"Reset command received for chat_id: {chat_id}")
    group_notes[chat_id] = {}
    group_keys[chat_id] = {"P": set(), "H": set()}
    sorted_notes[chat_id] = {"P": {}, "H": {}}
    start_time_data.pop(chat_id, None)
    chat_data.pop(chat_id, None)
    await update.message.reply_text('تم إعادة تعيين الملاحظات إلى الصفر.')

async def collect_photos(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    message = update.message or update.edited_message
    if not message or not message.photo:
        logger.debug(f"No photo found in message for chat_id: {message.chat_id}")
        return

    chat_id = message.chat_id
    text = message.caption if message.caption else ""
    logger.debug(f"Photo received in chat_id: {chat_id}, caption: {text}")

    if text and not any(word in text for word in ignore_words):
        ignore_patterns = [
            r'\b\d+l\b',
            r'\b\d+L\b',
            r'\b\d+D\b',
            r'\b\d+d\b',
            r'\b\d+\*\d+\b',
            r'\b\d+x\d+\b'
        ]

        for pattern in ignore_patterns:
            text = re.sub(pattern, '', text)

        lines = text.splitlines()
        filtered_lines = [line for line in lines if line.strip()]
        if filtered_lines:
            update_notes(chat_id, "\n".join(filtered_lines), message.message_id)
            logger.debug(f"Notes updated for chat_id: {chat_id}, message_id: {message.message_id}")

            # معالجة الملاحظات لفرزها تحت P و H
            if chat_id not in sorted_notes:
                sorted_notes[chat_id] = {"P": {}, "H": {}}

            # إزالة الملاحظات القديمة المرتبطة بـ message_id
            for key_type in ["P", "H"]:
                for note in list(sorted_notes[chat_id][key_type]):
                    for key_id in list(sorted_notes[chat_id][key_type][note]):
                        if message.message_id in sorted_notes[chat_id][key_type][note][key_id]:
                            sorted_notes[chat_id][key_type][note][key_id].discard(message.message_id)
                            if not sorted_notes[chat_id][key_type][note][key_id]:
                                del sorted_notes[chat_id][key_type][note][key_id]
                            if not sorted_notes[chat_id][key_type][note]:
                                del sorted_notes[chat_id][key_type][note]

            current_key = None
            current_type = None
            for line in filtered_lines:
                line = line.strip()
                if re.match(r'^[Pp]\d+$', line, re.IGNORECASE):
                    current_key = int(re.search(r'\d+', line).group())
                    current_type = "P"
                elif re.match(r'^[Hh]\d+$', line, re.IGNORECASE):
                    current_key = int(re.search(r'\d+', line).group())
                    current_type = "H"
                elif current_type and line:
                    if line not in sorted_notes[chat_id][current_type]:
                        sorted_notes[chat_id][current_type][line] = {}
                    if current_key not in sorted_notes[chat_id][current_type][line]:
                        sorted_notes[chat_id][current_type][line][current_key] = set()
                    sorted_notes[chat_id][current_type][line][current_key].add(message.message_id)
        
        if chat_id not in start_time_data:
            start_time_data[chat_id] = message.date.astimezone(pytz.timezone('Asia/Baghdad'))
            logger.debug(f"Set start_time for chat_id: {chat_id}")

async def auto_clear(chat_id, context):
    logger.debug(f"Auto-clear triggered for chat_id: {chat_id}")
    await asyncio.sleep(16 * 60 * 60)  # الانتظار لمدة 16 ساعة
    if chat_id in group_notes:
        group_notes[chat_id] = {}
        group_keys[chat_id] = {"P": set(), "H": set()}
        start_time_data.pop(chat_id, None)
        sorted_notes[chat_id] = {"P": {}, "H": {}}
        logger.debug(f"Cleared group_notes, group_keys, and sorted_notes for chat_id: {chat_id}")
    if chat_id in chat_data:
        chat_data[chat_id] = {
            'hh_data': {},
            'p_data': {},
            'last_update': None,
            'used_hh_ids': set(),
            'used_closures': set(),
            'used_p_ids': set()
        }
        logger.debug(f"Cleared chat_data for chat_id: {chat_id}")

async def send_totals(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    chat_id = update.message.chat_id
    logger.debug(f"Send totals command received for chat_id: {chat_id}")
    group_name = update.message.chat.title if update.message.chat.title else "اسم المجموعة غير معروف"

    if chat_id not in group_keys:
        logger.debug(f"No group_keys found for chat_id: {chat_id}")
        await update.message.reply_text('لا توجد إحصائيات لهذه المجموعة.')
        return

    p_keys = sorted(group_keys[chat_id]["P"], key=lambda k: int(re.search(r'\d+', k).group()))
    h_keys = sorted(group_keys[chat_id]["H"], key=lambda k: int(re.search(r'\d+', k).group()))

    count_p = len(p_keys)
    count_h = len(h_keys)

    total_message = f"إحصائيات المجموعة: {group_name}\n\nالإجمالي:\nP: {count_p}\n" + "\n".join(p_keys) + f"\n\nH: {count_h}\n" + "\n".join(h_keys)
    await update.message.reply_text(total_message)

    context.application.create_task(auto_clear(chat_id, context))

async def send_notes(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    chat_id = update.message.chat_id
    logger.debug(f"Send notes command received for chat_id: {chat_id}")
    group_name = update.message.chat.title if update.message.chat.title else "اسم المجموعة غير معروف"

    if chat_id not in group_notes or not group_notes[chat_id]:
        logger.debug(f"No notes found for chat_id: {chat_id}")
        await update.message.reply_text('لا توجد ملاحظات لإرسالها.')
        return

    if chat_id in start_time_data:
        start_time = start_time_data[chat_id]
        current_time = datetime.now(pytz.timezone('Asia/Baghdad'))
        formatted_time = f"From - {start_time.strftime('%Y-%m-%d (%I:%M %p)')} \nTo      - {current_time.strftime('%Y-%m-%d (%I:%M %p')}"
    else:
        formatted_time = "N/A"

    header = f"ملاحظات المجموعة: {group_name}\n\nمجموع الوقت المستغرق:\n{formatted_time}\n\n"

    blocks = []

    for key_type in ["P", "H"]:
        keys = sorted(group_keys[chat_id][key_type], key=lambda k: int(re.search(r'\d+', k).group()))

        for key in keys:
            if key in group_notes[chat_id]:
                notes_content = "\n".join(dict.fromkeys(group_notes[chat_id][key].values()))
                unique_notes = "\n".join(line for line in dict.fromkeys(notes_content.splitlines()) if line.strip().upper() != key.upper())

                if re.search(r'\b(clear|c)\b', unique_notes, re.IGNORECASE):
                    continue

                if unique_notes.strip():
                    block = f"{key}:\n{unique_notes}\n\n"
                    blocks.append(block)

    if "NOTES" in group_notes[chat_id]:
        all_notes = [note for note in dict.fromkeys(group_notes[chat_id]["NOTES"].values()) if isinstance(note, str) and not re.search(r'\b(clear|c)\b', note, re.IGNORECASE)]
        if all_notes:
            unique_general_notes = "\n".join(dict.fromkeys(all_notes))
            block = "ملاحظات عامة:\n" + unique_general_notes + "\n"
            blocks.append(block)

    if not blocks:
        await update.message.reply_text('لا توجد ملاحظات لإرسالها.')
        return

    notes_content = ''.join(blocks)
    full_message = header + notes_content

    max_length = 4096
    if len(full_message) <= max_length:
        await update.message.reply_text(full_message)
    else:
        messages = []
        current_message = header
        for block in blocks:
            if len(current_message) + len(block) > max_length:
                messages.append(current_message)
                current_message = ""
            current_message += block
        if current_message:
            messages.append(current_message)

        for msg in messages:
            await update.message.reply_text(msg)

    context.application.create_task(auto_clear(chat_id, context))

# دالة فرز الملاحظات مع العنوان المطلوب
async def sort_note(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    chat_id = update.message.chat_id
    logger.debug(f"Sort note command received for chat_id: {chat_id}")
    group_name = update.message.chat.title if update.message.chat.title else "اسم المجموعة غير معروف"

    if chat_id not in sorted_notes or not (sorted_notes[chat_id]["P"] or sorted_notes[chat_id]["H"]):
        logger.debug(f"No sorted notes found for chat_id: {chat_id}")
        await update.message.reply_text('لا توجد ملاحظات لفرزها.')
        return

    if chat_id in start_time_data:
        start_time = start_time_data[chat_id]
        current_time = datetime.now(pytz.timezone('Asia/Baghdad'))
        formatted_time = f"From - {start_time.strftime('%Y-%m-%d (%I:%M %p)')} \nTo      - {current_time.strftime('%Y-%m-%d (%I:%M %p')}"
    else:
        formatted_time = "N/A"

    header = f"ملاحظات المجموعة المفرزة: {group_name}\n\nمجموع الوقت المستغرق:\n{formatted_time}\n\n"

    blocks = []

    for key_type in ["P", "H"]:
        block = f"{'ملاحظات الأعمدة Poles' if key_type == 'P' else 'هاندهولات HandHoles'}:\n\n"
        notes_data_text = []
        notes_data_numeric = []
        for note, keys in sorted_notes[chat_id][key_type].items():
            key_list = sorted(keys.keys())
            count = len(key_list)
            key_str = ",".join(str(k) for k in key_list)
            if re.match(r'^\d+\.\d+$', note):
                try:
                    numeric_value = float(note)
                    notes_data_numeric.append((note, key_str, count, numeric_value))
                except ValueError:
                    notes_data_text.append((note, key_str, count))
            else:
                notes_data_text.append((note, key_str, count))
        
        notes_data_text.sort(key=lambda x: (-x[2], x[0]))
        notes_data_numeric.sort(key=lambda x: (-x[2], x[3]))
        
        for note, key_str, count in notes_data_text:
            block += f"{note}\n{key_type}({key_str})\n{count} مرات\n\n"
        for note, key_str, count, _ in notes_data_numeric:
            block += f"{note}\n{key_type}({key_str})\n{count} مرات\n\n"
        
        if notes_data_text or notes_data_numeric:
            blocks.append(block)

    if not blocks:
        await update.message.reply_text('لا توجد ملاحظات لفرزها.')
        return

    full_message = header + "".join(blocks)

    max_length = 4096
    if len(full_message) <= max_length:
        await update.message.reply_text(full_message)
    else:
        messages = []
        current_message = header
        for block in blocks:
            if len(current_message) + len(block) > max_length:
                messages.append(current_message)
                current_message = ""
            current_message += block
        if current_message:
            messages.append(current_message)

        for msg in messages:
            await update.message.reply_text(msg)

    context.application.create_task(auto_clear(chat_id, context))

async def welcome_message(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    logger.debug(f"Welcome message triggered for chat_id: {update.message.chat_id}")
    for new_member in update.message.new_chat_members:
        if new_member.id == context.bot.id:
            group_name = update.message.chat.title if update.message.chat.title else "اسم المجموعة غير معروف"
            group_id = update.message.chat_id
            added_by = update.message.from_user
            added_by_name = added_by.username if added_by.username else f"{added_by.first_name} {added_by.last_name or ''}".strip() or "مجهول"
            added_by_id = added_by.id
            added_by_phone = ""

            try:
                logger.debug(f"Sending notification to YOUR_USER_ID: {YOUR_USER_ID}")
                await context.bot.send_message(
                    chat_id=YOUR_USER_ID,
                    text=f"تم إضافة البوت إلى مجموعة جديدة:\nاسم المجموعة: {group_name}\nمعرف المجموعة: {group_id}\nأُضيف بواسطة: {added_by_name} (ID: {added_by_id})\nرقم الهاتف: {added_by_phone if added_by_phone else 'غير متاح'}"
                )
            except Exception as e:
                logger.error(f"Error sending welcome notification: {str(e)}")
            
            welcome_text = """مرحباً انا @FTTHQCBOT  وظيفتي هي مساعدتكم في جمع الملاحظات، متابعة الإحصائيات، وإدارة بيانات HandHoles و Poles بسهولة. يجب إضافة البوت قبل البدء بعملية إرسال الصور أو النصوص.

📌 فوائد البوت:
- جمع الملاحظات المرسلة مع الصور.
- حساب الإحصائيات المتعلقة بالمجموعة.
- إدارة الملاحظات وتنظيمها بشكل تلقائي.
- عمل ملف شيت اكسل HandHoles و Poles  بنائاً على رسائل المرسلة بدون صور.

⚙️ أوامر البوت:
- /start - تفعيل البوت.
- /stop - إيقاف البوت وإعادة تعيين الملاحظات.
- /clear - إعادة تعيين جميع الملاحظات.
- /total - عرض الإحصائيات.
- /sendnote - إرسال جميع الملاحظات.
- /sortnote - فرز الملاحظات حسب Poles و HandHoles.
- /exportHH - تصدير بيانات HandHoles إلى Excel.
- /exportP - تصدير بيانات Poles إلى Excel.

🚀 كيفية الاستخدام:
1. أرسل صورة مع تسمية توضيحية (Caption) تحتوي على الملاحظات.
2. أرسل رسائل نصية تحتوي على بيانات HandHoles أو Poles بالتنسيق المحدد.
3. استخدم الأوامر المذكورة أعلاه لإدارة الملاحظات، الإحصائيات، وبيانات HandHoles و Poles.
4. سيتم مسح الملاحظات وبيانات HandHoles و Poles تلقائيًا بعد مرور ساعتين.

ملاحظة هامة‼️‼️
التنسيق المطلوب لعمل شيت اكسل خاص ب HandHoles ارسل كلاتي مثلاً
بنسبة الى كلوجر
H1 C1
48F1 in out
6l
25

36F6 out
8l
28

24F3 pass
8l
27

بنسبة الى باس ثرو
H2
24F1
7l
25

36F6
8l
25

بنسبة الى FDH HH, FDT HH
FDH HH
48F1 
7l
25

48F2
8l
25

48F3
7l
29

التنسيق المطلوب لعمل شيت اكسل خاص ب Poles ارسل كلاتي مثلاً
P1
36F6
6l
26

بعد انتهاء من ارسال جميع البيانات الخاصة ب HandHoles و Poles
فقط ارسل امر /exportHH للهاندهول و /exportP للأعمدة
ملاحظة‼️‼️
لا يهم ان كانت حروف صغيرة او كبيرة

Powered by @SajadFTTH ⚠️
"""
            await update.message.reply_text(welcome_text)

def calculate_loop_length(diameter_cm, loops):
    logger.debug(f"Calculating loop length with diameter: {diameter_cm}, loops: {loops}")
    return (diameter_cm * loops * 3.14) / 100

async def handle_message(update: Update, context):
    message = update.message or update.edited_message
    if not message or not message.text:
        logger.debug(f"No text found in message for chat_id: {message.chat_id}")
        return

    chat_id = message.chat.id
    message_id = message.message_id
    message_text = message.text.strip()
    chat_title = message.chat.title or "test"
    update_time = message.date
    logger.debug(f"Handling message in chat_id: {chat_id}, message_id: {message_id}, text: {message_text}")

    if 'f' not in message_text.lower():
        logger.debug("No 'f' in message, ignoring as notes list")
        return

    lines = [line.strip() for line in message_text.split('\n') if line.strip()]
    if not lines:
        return

    first_line = lines[0]
    if not re.match(r'^(h|H)\d+(?:\s+(c|C)\d+)?$|^(FDH|FDT)\s+HH$|^(p|P)\d+$', first_line, re.IGNORECASE):
        logger.debug(f"Invalid format for first line: {first_line}")
        return

    if chat_id not in chat_data:
        chat_data[chat_id] = {
            'hh_data': {},
            'p_data': {},
            'last_update': None,
            'used_hh_ids': set(),
            'used_closures': set(),
            'used_p_ids': set()
        }
        logger.debug(f"Initialized chat_data for chat_id: {chat_id}")

    chat_data[chat_id]['last_update'] = update_time

    is_update = update.edited_message is not None
    old_hh = None
    old_closure = None
    old_p = None
    old_type = None
    if is_update:
        if message_id in chat_data[chat_id]['hh_data']:
            old_type = 'hh'
            old_data = chat_data[chat_id]['hh_data'][message_id][0]
            old_hh = old_data['H.H ID']
            old_closure = old_data['NO.OF CLOSURE']
            del chat_data[chat_id]['hh_data'][message_id]
            chat_data[chat_id]['used_hh_ids'].discard(old_hh)
            if old_closure:
                chat_data[chat_id]['used_closures'].discard(old_closure)
        elif message_id in chat_data[chat_id]['p_data']:
            old_type = 'p'
            old_data = chat_data[chat_id]['p_data'][message_id][0]
            old_p = old_data['P ID']
            del chat_data[chat_id]['p_data'][message_id]
            chat_data[chat_id]['used_p_ids'].discard(old_p)

    current_hh = None
    current_closure = None
    current_p = None
    line_index = -1
    temp_data = []
    has_error = False
    data_type = None

    for i, line in enumerate(lines):
        line_index = i
        line = line.strip()

        if re.match(r'^(h|H)\d+(?:\s+(c|C)\d+)?$|^(FDH|FDT)\s+HH$', line, re.IGNORECASE):
            data_type = 'hh'
            parts = line.split()
            if len(parts) == 2 and (parts[0].lower().startswith('h') and parts[1].lower().startswith('c')):
                current_hh = parts[0].upper()
                current_closure = parts[1].upper()
                if current_hh in chat_data[chat_id]['used_hh_ids']:
                    await message.reply_text(f"H.H ID '{current_hh}' تم ذكره سابقًا.")
                    has_error = True
                if current_closure in chat_data[chat_id]['used_closures']:
                    await message.reply_text(f"NO.OF CLOSURE '{current_closure}' تم ذكره سابقًا.")
                    has_error = True
            elif len(parts) == 2 and parts[0].lower() in ['fdh', 'fdt'] and parts[1].lower() == 'hh':
                current_hh = f"{parts[0].upper()} HH"
                current_closure = None
                if current_hh in chat_data[chat_id]['used_hh_ids']:
                    await message.reply_text(f"H.H ID '{current_hh}' تم ذكره سابقًا.")
                    has_error = True
            elif parts[0].lower().startswith('h') and len(parts) == 1:
                current_hh = parts[0].upper()
                current_closure = None
                if current_hh in chat_data[chat_id]['used_hh_ids']:
                    await message.reply_text(f"H.H ID '{current_hh}' تم ذكره سابقًا.")
                    has_error = True
            continue
        elif re.match(r'^(p|P)\d+$', line, re.IGNORECASE):
            data_type = 'p'
            current_p = line.upper()
            if current_p in chat_data[chat_id]['used_p_ids']:
                await message.reply_text(f"P ID '{current_p}' تم ذكره سابقًا.")
                has_error = True
            continue
        elif line.startswith(('h', 'H', 'FDH', 'FDT', 'p', 'P')) and not has_error:
            await message.reply_text("تنسيق ID غير صحيح. يجب أن يكون على شكل H[رقم] [C[رقم] اختياري] أو FDH HH أو FDT HH أو P[رقم].")
            has_error = True
            continue

        if 'f' in line.lower() and not has_error:
            try:
                line_lower = line.lower()
                cable_type_match = re.match(r'^\d+', line_lower)
                if not cable_type_match:
                    raise ValueError("تنسيق Cable Type غير صحيح. يجب أن يبدأ بأرقام متبوعة بـ 'f'.")
                cable_type = cable_type_match.group(0)

                fiber_parts = line_lower.split('f')
                fiber = 'F' + ''.join(filter(str.isdigit, fiber_parts[1].split()[0])) if len(fiber_parts) > 1 else 'F1'

                if data_type == 'hh':
                    status_list = ['PASS']
                    if current_hh and current_hh.lower() in ['fdh hh', 'fdt hh']:
                        status_list = ['OUT']
                    elif any(x in line_lower for x in ['in', 'out', 'pass']):
                        if 'in out' in line_lower:
                            status_list = ['IN', 'OUT']
                        elif 'in' in line_lower:
                            status_list = ['IN']
                        elif 'out' in line_lower:
                            status_list = ['OUT']
                        elif 'pass' in line_lower:
                            status_list = ['PASS']
                    if not current_hh:
                        current_hh = 'H2'
                        status_list = ['PASS']
                else:
                    if not current_p:
                        await message.reply_text("لم يتم تحديد P ID قبل البيانات.")
                        has_error = True
                        continue
                    status_list = []

                loops = 0
                diameter = 0
                if line_index + 1 < len(lines):
                    next_line = lines[line_index + 1].strip()
                    if next_line.lower().endswith('l'):
                        loops = int(''.join(filter(str.isdigit, next_line)))
                        if line_index + 2 < len(lines):
                            diameter_line = lines[line_index + 2].strip()
                            if diameter_line.isdigit():
                                diameter = int(diameter_line)
                            else:
                                raise ValueError("قطر الكابل يجب أن يكون رقمًا صحيحًا.")
                    else:
                        raise ValueError("عدد الحلقات يجب أن ينتهي بـ 'l'.")
                else:
                    raise ValueError("بيانات غير كاملة: يجب توفير عدد الحلقات وقطر الكابل.")

                loop_length = calculate_loop_length(diameter, loops)

                if data_type == 'hh':
                    for status in status_list:
                        data_entry = {
                            'H.H ID': current_hh,
                            'NO.OF CLOSURE': current_closure,
                            'Cable Type': cable_type,
                            'NO.OF (F)': fiber,
                            'IN/OUT /PASS': status,
                            'Cable Loop Diameter (cm)': diameter,
                            'No. of loops': loops,
                            'Actual cable loop length (m)': loop_length,
                            'message_id': message_id
                        }
                        temp_data.append(data_entry)
                elif data_type == 'p':
                    data_entry = {
                        'P ID': current_p,
                        'Cable Type': cable_type,
                        'NO.OF (F)': fiber,
                        'Cable Loop Diameter (cm)': diameter,
                        'No. of loops': loops,
                        'Actual cable loop length (m)': loop_length,
                        'message_id': message_id
                    }
                    temp_data.append(data_entry)
            except (ValueError, IndexError) as e:
                await message.reply_text(f"خطأ في تنسيق البيانات: {str(e)}. الرجاء التحقق من البيانات وتعديلها.")
                has_error = True
                continue

    if temp_data and not has_error:
        try:
            if data_type == 'hh':
                chat_data[chat_id]['hh_data'][message_id] = temp_data
                chat_data[chat_id]['used_hh_ids'].add(current_hh)
                if current_closure:
                    chat_data[chat_id]['used_closures'].add(current_closure)
            elif data_type == 'p':
                chat_data[chat_id]['p_data'][message_id] = temp_data
                chat_data[chat_id]['used_p_ids'].add(current_p)
            logger.info(f"{'Updated' if is_update else 'Added'} data with message_id: {message_id}")
        except Exception as e:
            logger.error(f"Error processing data for message_id {message_id}: {str(e)}")
            await message.reply_text("حدث خطأ داخلي أثناء معالجة البيانات، حاول مرة أخرى لاحقًا.")
    elif is_update and has_error:
        pass
    elif has_error:
        await message.reply_text("تم تجاهل الإدخال بسبب أخطاء في التنسيق أو تكرار القيم. استخدم /exportHH أو /exportP لتصدير البيانات الصحيحة السابقة إذا وجدت.")

async def export_excel(update: Update, context, data_type='hh'):
    chat_id = update.effective_chat.id
    logger.debug(f"Export Excel triggered for chat_id: {chat_id}, data_type: {data_type}")
    data_key = 'hh_data' if data_type == 'hh' else 'p_data'
    if chat_id not in chat_data or not any(chat_data[chat_id][data_key].values()):
        logger.debug(f"No data found for {data_type} in chat_id: {chat_id}")
        await update.message.reply_text(f"لا توجد بيانات {'HH' if data_type == 'hh' else 'P'} للتصدير. من فضلك أرسل البيانات أولاً.")
        return

    chat_title = update.effective_chat.title or "test"
    all_data = []

    for message_id in chat_data[chat_id][data_key]:
        all_data.extend(chat_data[chat_id][data_key][message_id])

    df = pd.DataFrame(all_data)

    if data_type == 'hh':
        df['sort_key'] = df['H.H ID'].apply(
            lambda x: (
                0 if x.lower() in ['fdh hh', 'fdt hh'] else 1,
                int(re.search(r'\d+', x).group(0)) if re.search(r'\d+', x) else 0
            )
        )
        df_sorted = df.sort_values(by='sort_key').drop(columns=['sort_key', 'message_id'])
    else:
        df['sort_key'] = df['P ID'].apply(
            lambda x: int(re.search(r'\d+', x).group(0)) if re.search(r'\d+', x) else 0
        )
        df_sorted = df.sort_values(by='sort_key').drop(columns=['sort_key', 'message_id'])

    wb = Workbook()
    ws = wb.active

    columns_count = 8 if data_type == 'hh' else 6
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=columns_count)
    header_cell = ws.cell(row=1, column=1, value=chat_title.upper())
    header_cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    header_cell.alignment = Alignment(horizontal='center', vertical='center')

    if data_type == 'hh':
        headers = ['H.H ID', 'NO.OF CLOSURE', 'Cable Type', 'NO.OF (F)', 'IN/OUT /PASS', 
                   'Cable Loop Diameter (cm)', 'No. of loops', 'Actual cable loop length (m)']
    else:
        headers = ['P ID', 'Cable Type', 'NO.OF (F)', 
                   'Cable Loop Diameter (cm)', 'No. of loops', 'Actual cable loop length (m)']

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.fill = PatternFill(
            start_color="FFC000" if col <= (5 if data_type == 'hh' else 3) else "8EAADB",
            end_color="FFC000" if col <= (5 if data_type == 'hh' else 3) else "8EAADB",
            fill_type="solid"
        )

    previous_hh_id = None
    previous_closure = None

    for row_idx, row in enumerate(df_sorted.itertuples(index=False), start=3):
        for col_idx, value in enumerate(row, 1):
            if data_type == 'hh':
                if col_idx == 1:
                    if value == previous_hh_id:
                        value = None
                    else:
                        previous_hh_id = value
                elif col_idx == 2:
                    if value == previous_closure:
                        value = None
                    else:
                        previous_closure = value
            else:
                if col_idx == 1:
                    if value == previous_hh_id:
                        value = None
                    else:
                        previous_hh_id = value
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if col_idx == columns_count:
                cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")

    for col_idx in range(1, columns_count + 1):
        max_length = 0
        column = chr(64 + col_idx)
        for cell in ws[column]:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except AttributeError:
                pass
        adjusted_width = max_length + 5
        ws.column_dimensions[column].width = adjusted_width

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    excel_filename = (f"{chat_title} - (Holes) LM Loop length sheet.xlsx" if data_type == 'hh' 
                      else f"{chat_title} - (Poles) LM Loop length sheet.xlsx")
    
    try:
        await context.bot.send_document(
            chat_id=chat_id,
            document=output,
            filename=excel_filename
        )
        if data_type == 'hh':
            await update.message.reply_text("تم تصدير ملف Excel لـ HandHoles.")
        else:
            await update.message.reply_text("تم تصدير ملف Excel لـ Poles.")
    except Exception as e:
        logger.error(f"Error sending Excel file: {str(e)}")
        await update.message.reply_text("حدث خطأ أثناء تصدير الملف، حاول مرة أخرى لاحقًا.")

async def export_hh(update: Update, context):
    await export_excel(update, context, data_type='hh')

async def export_p(update: Update, context):
    await export_excel(update, context, data_type='p')

async def notify_usage(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    group_name = update.message.chat.title if update.message.chat.title else "اسم المجموعة غير معروف"
    group_id = update.message.chat_id
    user = update.message.from_user
    username = user.username if user.username else user.first_name
    user_id = user.id
    user_phone = ""
    logger.debug(f"Command used by {username} (ID: {user.id}) in chat_id: {group_id}")

    try:
        await context.bot.send_message(
            chat_id=YOUR_USER_ID,
            text=f"تم استخدام البوت في المجموعة:\nاسم المجموعة: {group_name}\nمعرف المجموعة: {group_id}\nالمستخدم: {username} (ID: {user_id})\nرقم الهاتف: {user_phone if user_phone else 'غير متاح'}"
        )
    except Exception as e:
        logger.error(f"Error sending usage notification: {str(e)}")

def wrap_command_handler(handler):
    async def wrapped(update: Update, context: ContextTypes.DEFAULT_TYPE):
        await notify_usage(update, context)
        await handler(update, context)
    return wrapped

if __name__ == "__main__":
    API_TOKEN = os.getenv("BOT_TOKEN")
    if not API_TOKEN:
        raise RuntimeError("BOT_TOKEN not found. Add it in Fly secrets.")
    
    application = ApplicationBuilder().token(API_TOKEN).build()

    application.add_handler(CommandHandler("start", wrap_command_handler(start)))
    application.add_handler(CommandHandler("stop", wrap_command_handler(stop)))
    application.add_handler(CommandHandler("clear", wrap_command_handler(reset)))
    application.add_handler(CommandHandler("total", wrap_command_handler(send_totals)))
    application.add_handler(CommandHandler("sendnote", wrap_command_handler(send_notes)))
    application.add_handler(CommandHandler("sortnote", wrap_command_handler(sort_note)))  # تغيير هنا
    application.add_handler(CommandHandler("exportHH", wrap_command_handler(export_hh)))
    application.add_handler(CommandHandler("exportP", wrap_command_handler(export_p)))
    application.add_handler(MessageHandler(filters.PHOTO, collect_photos))
    application.add_handler(MessageHandler(filters.StatusUpdate.NEW_CHAT_MEMBERS, welcome_message))
    application.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_message))

    async def error_handler(update: Update, context):
        logger.error(f"Occurred error: {str(context.error)} with update: {update}")
        if update and update.effective_message:
            await update.effective_message.reply_text("حدث خطأ غير متوقع، حاول مرة أخرى لاحقًا.")
    
    application.add_error_handler(error_handler)
    application.run_polling(allowed_updates=Update.ALL_TYPES)
